import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors


match_fill = PatternFill(patternType="solid", fgColor=colors.Color(rgb="0000FF00"))

def find_matches(file1, file2, column1, column2, out_file):
    '''
    Find matches and save highlighted cells to copy of second sheet
    '''
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    wb1_active = wb1.active
    wb2_active = wb2.active

    x=0

    for rowA in wb1_active.rows:
        for rowB in wb2_active.rows:
            if rowA[column1].value == rowB[column2].value:
                rowB[column2].fill = match_fill
                x+=1

    print("{} matches found".format(x))
    wb2.save(filename=out_file)


print("PY_XL Match Finder")
print("(Explicit paths preferred)")
file1 = input("First xlsx: ")
column1 = int(input("Column to search: "))
file2 = input("Second xlsx: ")
column2 = int(input("Column to search: "))
out_file = input("Output file: ")

find_matches(file1, file2, column1, column2, out_file)
